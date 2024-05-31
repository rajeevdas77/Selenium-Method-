import openpyxl
#flie--->workbook----->sheets----->rows---->cells
file="D:\Selenium Automation\screenshot\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["sheet1"]
row=sheet.max_row
cols=sheet.max_column
for r in range(1,row+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='       ')
    print()    